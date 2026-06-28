import io
import re
import unittest
from pathlib import Path

from fastapi.testclient import TestClient
from jinja2 import Environment
from openpyxl import load_workbook
from pypdf import PdfReader
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.database import Base, get_db
from app.i18n import PHRASES_EN, localized_name, translate_message, translate_text, translate_value
from app.main import app
from app.models.core import Category, Department, ProcurementCase, Product, Requisition, RequisitionItem, Role, User
from app.security import PERMISSIONS, hash_password
from app.services.exports import rows_to_xlsx
from app.services.notifications import localize_notification
from app.services.procurement_pdf import procurement_form_to_pdf
from app.services.requisition_pdf import requisition_to_pdf
from app.services.tdr_pdf import terms_of_reference_to_pdf


ROOT = Path(__file__).resolve().parents[1]


class TranslationCatalogTests(unittest.TestCase):
    def test_all_template_literals_have_english_translations_and_parse(self):
        missing = set()
        environment = Environment()
        for path in (ROOT / "app" / "templates").rglob("*.html"):
            source = path.read_text(encoding="utf-8")
            environment.parse(source)
            for match in re.finditer(r"""\b(?:tx|tm)\(\s*(["'])(.*?)\1""", source, re.DOTALL):
                if match.group(2) not in PHRASES_EN:
                    missing.add(match.group(2))
        self.assertEqual(missing, set())

    def test_permission_labels_and_common_dynamic_messages_are_translated(self):
        self.assertTrue(all(label in PHRASES_EN for label in PERMISSIONS.values()))
        self.assertEqual(translate_message("Departamento é obrigatório.", "en"), "Department is required.")
        self.assertEqual(translate_message("Quantidade deve ser um número válido.", "en"), "Quantity must be a valid number.")
        self.assertEqual(translate_message("O item AC-001 - Ar condicionado não tem stock disponível.", "en"), "Item AC-001 - Ar condicionado has no available stock.")
        self.assertEqual(translate_value("Pending Budget Verification", "pt"), "Aguarda verificação orçamental")
        self.assertEqual(translate_value("Pending Budget Verification", "en"), "Pending budget verification")

    def test_xlsx_metadata_and_headers_follow_language(self):
        content = rows_to_xlsx(["Code", "Product"], [["A-1", "Original user data"]], "Stock Report", "en")
        sheet = load_workbook(io.BytesIO(content)).active
        self.assertEqual(sheet["A1"].value, "Stock Management System")
        self.assertTrue(sheet["B2"].value.startswith("Generated on "))
        self.assertEqual(sheet["A4"].value, "Code")
        self.assertEqual(sheet["B5"].value, "Original user data")

    def test_all_stock_report_headers_are_translated(self):
        expected = {
            "Código": "Code",
            "Produto": "Product",
            "Categoria": "Category",
            "Unidade": "Unit",
            "Preço Unit.": "Unit price",
            "Stock Atual": "Current stock",
            "Stock Mínimo": "Minimum stock",
            "Entradas": "Entries",
            "Saídas": "Issues",
            "Estado": "Status",
            "Monitorizado": "Monitored",
            "Alerta": "Alert",
            "Fornecedor": "Supplier",
            "Valor PO": "PO value",
        }
        self.assertEqual(
            {source: translate_text(source, "en") for source in expected},
            expected,
        )


class I18nApplicationTests(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite:///:memory:",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.SessionLocal = sessionmaker(bind=self.engine, autoflush=False, autocommit=False)
        self.db = self.SessionLocal()
        role = Role(name="SuperAdmin")
        department = Department(name="Operações")
        category = Category(name="Climatização", name_en="Air Conditioning", normalized_name="climatizacao")
        self.db.add_all([role, department, category])
        self.db.flush()
        self.user = User(
            full_name="Administrador Principal",
            username="superadmin",
            email="superadmin@example.com",
            password_hash=hash_password("Admin@12345"),
            role_id=role.id,
            department_id=department.id,
            preferred_language="pt",
        )
        self.db.add(self.user)
        self.db.flush()
        self.product = Product(
            code="AC-001",
            name="Ar condicionado de teste",
            name_en="Test air conditioner",
            category_id=category.id,
            unit="un",
            unit_price=100,
            minimum_stock=2,
            current_stock=5,
            created_by_id=self.user.id,
        )
        self.db.add(self.product)
        self.db.commit()
        app.dependency_overrides[get_db] = self.override_db
        self.client = TestClient(app)
        response = self.client.post(
            "/login",
            data={"username": "superadmin", "password": "Admin@12345"},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)

    def tearDown(self):
        app.dependency_overrides.clear()
        self.client.close()
        self.db.close()
        self.engine.dispose()

    def override_db(self):
        db = self.SessionLocal()
        try:
            yield db
        finally:
            db.close()

    def switch_language(self, language):
        response = self.client.post(
            "/preferencias/idioma",
            data={"language": language},
            headers={"referer": "http://testserver/dashboard"},
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)

    def test_english_renders_every_main_module_and_preserves_user_data(self):
        self.switch_language("en")
        paths = [
            "/dashboard",
            "/produtos",
            "/produtos/novo",
            "/movimentos",
            "/movimentos/novo",
            "/requisicoes",
            "/requisicoes/nova",
            "/procurement",
            "/procurement/nova",
            "/procurement/reposicao/nova",
            "/relatorios",
            "/relatorios/stock",
            "/relatorios/movimentos",
            "/relatorios/requisicoes",
            "/relatorios/procurement",
            "/relatorios/critico",
            "/utilizadores",
            "/perfis",
            "/configuracoes",
            "/configuracoes/matriz",
            "/notificacoes",
            "/auditoria",
            "/importar",
            "/documentos",
            "/sobre",
        ]
        pages = {}
        for path in paths:
            response = self.client.get(path)
            self.assertEqual(response.status_code, 200, path)
            self.assertIn('<html lang="en">', response.text, path)
            pages[path] = response.text
        self.assertIn("Test air conditioner", pages["/produtos"])
        self.assertNotIn(">Ar condicionado de teste<", pages["/produtos"])
        self.assertIn("Air Conditioning", pages["/produtos"])
        self.assertIn("Users", pages["/utilizadores"])
        self.assertIn("Approval matrix", pages["/configuracoes/matriz"])
        self.assertIn("Audit trail", pages["/auditoria"])

    def test_portuguese_renders_accents_and_original_business_names(self):
        response = self.client.get("/produtos")
        self.assertEqual(response.status_code, 200)
        self.assertIn('<html lang="pt">', response.text)
        self.assertIn("Ar condicionado de teste", response.text)
        self.assertIn("Climatização", response.text)
        self.assertIn("Configurações", response.text)
        self.assertNotIn("RequisiÃ", response.text)

    def test_language_persists_after_logout_and_login(self):
        self.switch_language("en")
        self.client.post("/logout", follow_redirects=False)
        login_page = self.client.get("/login")
        self.assertIn('<html lang="en">', login_page.text)
        self.assertIn("Username", login_page.text)
        self.client.post(
            "/login",
            data={"username": "superadmin", "password": "Admin@12345"},
            follow_redirects=False,
        )
        response = self.client.get("/dashboard")
        self.assertIn('<html lang="en">', response.text)
        with self.SessionLocal() as db:
            self.assertEqual(db.get(User, self.user.id).preferred_language, "en")

    def test_report_exports_follow_selected_language(self):
        self.switch_language("en")
        xlsx = self.client.get("/relatorios/stock?export=xlsx")
        self.assertEqual(xlsx.status_code, 200)
        sheet = load_workbook(io.BytesIO(xlsx.content)).active
        self.assertEqual(sheet["A1"].value, "Stock Management System")
        self.assertEqual(sheet["B4"].value, "Product")
        self.assertEqual(sheet["B5"].value, "Test air conditioner")
        self.assertEqual(sheet["C5"].value, "Air Conditioning")

        pdf = self.client.get("/relatorios/stock?export=pdf")
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf.headers["content-type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF"))

    def test_requisition_procurement_and_tdr_pdfs_follow_language(self):
        requisition = Requisition(
            number="REQ-2026-001",
            requesting_user_id=self.user.id,
            department_id=self.user.department_id,
            authorization_person="Chefe do Terminal",
            estimated_value=100,
            req_type="REQUISIÇÃO",
            status="Approved",
        )
        requisition.items.append(
            RequisitionItem(
                product_id=self.product.id,
                quantity_requested=1,
                quantity_issued=1,
                quantity_rejected=0,
                review_status="Aprovado",
            )
        )
        procurement = ProcurementCase(
            requisition=requisition,
            description="Manutenção preventiva do terminal",
            priority="Normal",
            item_type="Serviço",
            estimated_budget=100,
            status="Pending Budget Verification",
            tor_status="Pending HOD Approval",
            hse_documents_status="Not Required",
        )
        self.db.add(procurement)
        self.db.commit()

        self.user.preferred_language = "en"
        english_documents = [
            requisition_to_pdf(requisition, self.user),
            procurement_form_to_pdf(procurement, self.user),
            terms_of_reference_to_pdf(procurement, self.user),
        ]
        english_text = [
            "\n".join(page.extract_text() or "" for page in PdfReader(io.BytesIO(content)).pages)
            for content in english_documents
        ]
        self.assertIn("REQUISITION", english_text[0])
        self.assertIn("Pending authorisation by", english_text[0])
        self.assertIn("Estimated value", english_text[0])
        self.assertIn("Non-Stock Requisition Form", english_text[1])
        self.assertIn("TERM OF REFERENCE", english_text[2])
        self.assertNotIn("TERMO DE REFERÊNCIA", english_text[2])

        self.user.preferred_language = "pt"
        portuguese_tdr = terms_of_reference_to_pdf(procurement, self.user)
        portuguese_text = "\n".join(
            page.extract_text() or "" for page in PdfReader(io.BytesIO(portuguese_tdr)).pages
        )
        self.assertIn("TERMO DE REFERÊNCIA", portuguese_text)
        self.assertNotIn("TERM OF REFERENCE", portuguese_text)

    def test_optional_department_is_accepted_and_invalid_fields_are_friendly(self):
        response = self.client.post(
            "/utilizadores/novo",
            data={
                "full_name": "User Without Department",
                "username": "nodepartment",
                "email": "",
                "phone": "",
                "role_id": str(self.user.role_id),
                "department_id": "",
                "preferred_language": "en",
                "password": "Password123",
            },
            follow_redirects=False,
        )
        self.assertEqual(response.status_code, 303)

        response = self.client.post(
            "/utilizadores/novo",
            data={
                "full_name": "Broken Role",
                "username": "brokenrole",
                "role_id": "",
                "department_id": "",
                "preferred_language": "en",
                "password": "Password123",
            },
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn("Profile is required.", response.text)
        self.assertNotIn("int_parsing", response.text)

    def test_notifications_are_localized_per_recipient(self):
        self.user.preferred_language = "en"
        self.assertEqual(
            localize_notification("TdR para aprovação HOD: NS-001", self.user),
            "ToR pending HOD approval: NS-001",
        )
        self.assertIn(
            "Requester:",
            localize_notification("Requisitante: Nome Original", self.user),
        )
        self.assertEqual(localized_name(self.product, self.user), "Test air conditioner")
