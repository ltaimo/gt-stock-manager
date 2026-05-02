import csv
import io
import json
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import xlrd
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.config import get_settings
from app.models.core import Category, Department, Product, Role, User
from app.security import hash_password
from app.services.inventory import StockError, post_movement


PREVIEW_DIR = get_settings().uploads_dir / "import_previews"
PREVIEW_DIR.mkdir(parents=True, exist_ok=True)


@dataclass
class ImportResult:
    batch_id: str
    imported: int
    failed: int
    errors: list[dict]


def normalize(value: Any) -> str:
    return " ".join(str(value or "").strip().split()).title()


def normalize_key(value: Any) -> str:
    return str(value or "").strip().casefold()


def as_text(row: dict, *keys: str) -> str:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return str(value).strip()
    return ""


def as_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return None


def first_present(row: dict, *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if value not in (None, ""):
            return value
    return None


def canonical_action(value: Any) -> str:
    action = str(value or "").strip().upper()
    return {
        "SAIDA": "SAÍDA",
        "SAÍDA": "SAÍDA",
        "DEVOLUCAO": "DEVOLUÇÃO",
        "DEVOLUÇÃO": "DEVOLUÇÃO",
    }.get(action, action)


def _rows_from_matrix(matrix: list[list]) -> list[dict]:
    header_idx = None
    for idx, row in enumerate(matrix[:30]):
        non_empty = [str(v).strip() for v in row if v not in ("", None)]
        if len(non_empty) >= 2:
            header_idx = idx
            break
    if header_idx is None:
        return []
    headers = [str(h or "").strip() for h in matrix[header_idx]]
    rows = []
    for offset, row in enumerate(matrix[header_idx + 1 :], start=header_idx + 2):
        if any(cell not in ("", None) for cell in row):
            record = dict(zip(headers, row))
            record["_row_number"] = offset
            rows.append(record)
    return rows


def parse_table(filename: str, content: bytes, sheet_name: str | None = None) -> list[dict]:
    if filename.lower().endswith(".csv"):
        text = content.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
        for idx, row in enumerate(rows, start=2):
            row["_row_number"] = idx
        return rows

    if filename.lower().endswith(".xls"):
        try:
            wb = xlrd.open_workbook(file_contents=content)
            ws = wb.sheet_by_name(sheet_name) if sheet_name and sheet_name in wb.sheet_names() else wb.sheet_by_index(0)
            matrix = [[ws.cell_value(row_idx, col) for col in range(ws.ncols)] for row_idx in range(ws.nrows)]
            return _rows_from_matrix(matrix)
        except xlrd.biffh.XLRDError:
            pass

    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    matrix = [list(row) for row in ws.iter_rows(values_only=True)]
    return _rows_from_matrix(matrix)


def parse_old_stock_manager(filename: str, content: bytes) -> dict[str, list[dict]]:
    if filename.lower().endswith(".csv"):
        rows = parse_table(filename, content)
        headers = {normalize_key(key) for key in rows[0].keys()} if rows else set()
        if {"codigo", "código", "produto/marca"} & headers:
            return {"products": rows, "users": [], "movements": []}
        if {"usuarios", "usuários", "nome"} <= headers:
            return {"products": [], "users": rows, "movements": []}
        return {"products": [], "users": [], "movements": rows}
    return {
        "products": parse_table(filename, content, "ECONOMATO"),
        "users": parse_table(filename, content, "USERS"),
        "movements": parse_table(filename, content, "MOVIMENTO"),
    }


def error(module: str, row: dict, message: str) -> dict:
    clean = {key: value for key, value in row.items() if key != "_row_number"}
    return {"module": module, "row": row.get("_row_number", ""), "error": message, "data": clean}


def build_import_preview(db: Session, filename: str, content: bytes) -> dict:
    parsed = parse_old_stock_manager(filename, content)
    errors: list[dict] = []
    warnings: list[dict] = []

    existing_codes = {code for (code,) in db.execute(select(Product.code)).all()}
    existing_usernames = {username.casefold() for (username,) in db.execute(select(User.username)).all()}
    product_codes_seen: set[str] = set()
    product_names_available: set[str] = {name.casefold() for (name,) in db.execute(select(Product.name)).all()}
    valid_products: list[dict] = []

    for row in parsed["products"]:
        row_has_error = False
        code = as_text(row, "Código", "Codigo", "code")
        name = as_text(row, "Produto/Marca", "Produto", "name")
        category = as_text(row, "Categoria", "category")
        stock = as_number(first_present(row, "Qtde", "Quantidade", "current_stock"))
        minimum = as_number(first_present(row, "Mínimo", "Minimo", "minimum_stock")) or 0
        total_entries = as_number(first_present(row, "ENTRADA", "Entradas", "entries"))
        total_exits = as_number(first_present(row, "SAÍDA", "SAIDA", "Saídas", "Saidas", "exits"))

        if not code:
            errors.append(error("Produtos", row, "Código do produto é obrigatório."))
            row_has_error = True
        if not name:
            errors.append(error("Produtos", row, "Nome do produto está vazio."))
            row_has_error = True
        if not category:
            errors.append(error("Produtos", row, "Categoria está vazia ou nula."))
            row_has_error = True
        if code and code in existing_codes:
            errors.append(error("Produtos", row, f"Código de produto já existe na base de dados: {code}."))
            row_has_error = True
        if code and code in product_codes_seen:
            errors.append(error("Produtos", row, f"Código de produto duplicado no ficheiro: {code}."))
            row_has_error = True
        if stock is None:
            errors.append(error("Produtos", row, "Stock atual inválido."))
            row_has_error = True
        elif stock < 0:
            errors.append(error("Produtos", row, "Stock negativo não pode ser importado automaticamente."))
            row_has_error = True
        if total_entries is not None and total_entries < 0:
            errors.append(error("Produtos", row, "Quantidade de ENTRADA inválida; não pode ser negativa."))
            row_has_error = True
        if total_exits is not None and total_exits < 0:
            errors.append(error("Produtos", row, "Quantidade de SAÍDA inválida; não pode ser negativa."))
            row_has_error = True

        product_codes_seen.add(code)
        if not row_has_error:
            valid_products.append(
                {
                    "code": code,
                    "name": name,
                    "category": normalize(category),
                    "unit": as_text(row, "Unidade", "unit") or "un",
                    "current_stock": stock,
                    "minimum_stock": minimum,
                    "row": row.get("_row_number", ""),
                }
            )
            product_names_available.add(name.casefold())

    usernames_seen: set[str] = set()
    valid_users: list[dict] = []
    for row in parsed["users"]:
        row_has_error = False
        username = as_text(row, "Usuários", "Usuarios", "username")
        name = as_text(row, "Nome", "full_name")
        role = as_text(row, "Tipo", "role") or "User"
        department = as_text(row, "Departamento", "department") or "Geral"

        if not username:
            errors.append(error("Utilizadores", row, "Username está vazio."))
            row_has_error = True
        if not name:
            errors.append(error("Utilizadores", row, "Nome do utilizador está vazio."))
            row_has_error = True
        if username and username.casefold() in existing_usernames:
            errors.append(error("Utilizadores", row, f"Username já existe na base de dados: {username}."))
            row_has_error = True
        if username and username.casefold() in usernames_seen:
            errors.append(error("Utilizadores", row, f"Username duplicado no ficheiro: {username}."))
            row_has_error = True
        usernames_seen.add(username.casefold())

        if not row_has_error:
            valid_users.append(
                {
                    "username": username,
                    "full_name": name,
                    "role": role if role in {"SuperAdmin", "Admin", "Editor", "User"} else "User",
                    "department": normalize(department),
                    "row": row.get("_row_number", ""),
                }
            )

    valid_movements: list[dict] = []
    for row in parsed["movements"]:
        row_has_error = False
        action = canonical_action(as_text(row, "Acção", "Accão", "Acao", "Accao", "ACÇÃO", "ACCÃO"))
        item_name = as_text(row, "Item", "Produto", "Produto/Marca")
        quantity = as_number(first_present(row, "Quantidade", "Qtde", "1"))
        destination = as_text(row, "Destino", "Departamento/Destino")

        if not action and not item_name:
            continue
        if action not in {"ENTRADA", "SAÍDA", "DEVOLUÇÃO", "ACERTO"}:
            errors.append(error("Movimentos", row, f"Tipo de acção inválido: {action or '(vazio)'}."))
            row_has_error = True
        if not item_name:
            errors.append(error("Movimentos", row, "Item do movimento está vazio."))
            row_has_error = True
        elif item_name.casefold() not in product_names_available:
            errors.append(error("Movimentos", row, f"Produto do movimento não existe no ficheiro nem na base de dados: {item_name}."))
            row_has_error = True
        if quantity is None or quantity <= 0:
            errors.append(error("Movimentos", row, "Quantidade inválida; deve ser superior a zero."))
            row_has_error = True

        if not row_has_error:
            valid_movements.append(
                {
                    "action": action,
                    "item_name": item_name,
                    "quantity": quantity,
                    "destination": destination,
                    "responsible": as_text(row, "Responsável", "Responsavel", "RESPONSAVEL"),
                    "reference": as_text(row, "TIPO", "Referencia", "Referência"),
                    "row": row.get("_row_number", ""),
                }
            )

    batch_id = str(uuid.uuid4())
    preview = {
        "batch_id": batch_id,
        "filename": filename,
        "created_at": datetime.utcnow().isoformat(),
        "counts": {
            "products_total": len(parsed["products"]),
            "users_total": len(parsed["users"]),
            "movements_total": len(parsed["movements"]),
            "products_valid": len(valid_products),
            "users_valid": len(valid_users),
            "movements_valid": len(valid_movements),
            "errors": len(errors),
            "warnings": len(warnings),
        },
        "products": valid_products,
        "users": valid_users,
        "movements": valid_movements,
        "errors": errors,
        "warnings": warnings,
    }
    save_preview(preview)
    return preview


def preview_path(batch_id: str) -> Path:
    return PREVIEW_DIR / f"{batch_id}.json"


def save_preview(preview: dict) -> None:
    preview_path(preview["batch_id"]).write_text(json.dumps(preview, ensure_ascii=False, default=str, indent=2), encoding="utf-8")


def load_preview(batch_id: str) -> dict:
    path = preview_path(batch_id)
    if not path.exists():
        raise FileNotFoundError(batch_id)
    return json.loads(path.read_text(encoding="utf-8"))


def import_preview(db: Session, preview: dict, actor: User) -> ImportResult:
    if preview["errors"]:
        return ImportResult(preview["batch_id"], 0, len(preview["errors"]), preview["errors"])

    imported = 0
    product_by_name = {p.name.casefold(): p for p in db.scalars(select(Product)).all()}
    for item in preview["products"]:
        category = db.scalar(select(Category).where(Category.normalized_name == item["category"].lower()))
        if not category:
            category = Category(name=item["category"], normalized_name=item["category"].lower())
            db.add(category)
            db.flush()
        product = Product(
            code=item["code"],
            name=item["name"],
            category_id=category.id,
            unit=item["unit"],
            current_stock=0,
            minimum_stock=item["minimum_stock"],
            created_by_id=actor.id,
        )
        db.add(product)
        db.flush()
        product_by_name[product.name.casefold()] = product
        if item["current_stock"] > 0:
            post_movement(
                db,
                product=product,
                action_type="ENTRADA",
                quantity=item["current_stock"],
                registered_by=actor,
                notes="Stock inicial importado do Excel legado",
                reference_number=f"IMPORT-{preview['batch_id'][:8]}",
            )
        imported += 1

    default_password = hash_password("Reset@12345")
    for item in preview["users"]:
        role = db.scalar(select(Role).where(Role.name == item["role"])) or db.scalar(select(Role).where(Role.name == "User"))
        department = db.scalar(select(Department).where(Department.name == item["department"]))
        if not department:
            department = Department(name=item["department"])
            db.add(department)
            db.flush()
        db.add(
            User(
                full_name=item["full_name"],
                username=item["username"],
                password_hash=default_password,
                role_id=role.id,
                department_id=department.id,
                must_reset_password=True,
            )
        )
        imported += 1

    for item in preview["movements"]:
        product = product_by_name.get(item["item_name"].casefold())
        if not product:
            raise StockError(f"Produto não encontrado ao confirmar movimento: {item['item_name']}")
        post_movement(
            db,
            product=product,
            action_type=item["action"],
            quantity=item["quantity"],
            registered_by=actor,
            destination=item["destination"],
            responsible_person=item["responsible"],
            notes="Movimento importado do Excel legado",
            reference_number=item["reference"] or f"IMPORT-{preview['batch_id'][:8]}",
        )
        imported += 1

    return ImportResult(preview["batch_id"], imported, 0, [])
